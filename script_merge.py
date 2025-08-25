import openpyxl
from docx import Document
from docx2pdf import convert
import os
from datetime import datetime

# Configuración
EXCEL_FILE = "base_datos.xlsx"
TEMPLATE_FILE = "plantilla_certificado.docx"
OUTPUT_DIR = "certificados"

# Crear carpeta de salida si no existe
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Cargar base de datos de Excel
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active

# Asumimos que la primera fila es el encabezado
headers = {cell.value: idx for idx, cell in enumerate(ws[1])}

for row in ws.iter_rows(min_row=2, values_only=True):
    numero_registro = row[headers["numero_registro"]]
    fecha = row[headers["fecha"]]
    # Si la fecha está en formato Excel, convertirla
    if isinstance(fecha, datetime):
        fecha_str = fecha.strftime("%d/%m/%Y")
    else:
        fecha_str = str(fecha)

    # Crear documento Word desde la plantilla
    doc = Document(TEMPLATE_FILE)
    for p in doc.paragraphs:
        if "{numero_registro}" in p.text or "{fecha}" in p.text:
            p.text = p.text.replace("{numero_registro}", str(numero_registro))
            p.text = p.text.replace("{fecha}", fecha_str)

    output_word = os.path.join(OUTPUT_DIR, f"certificado_{numero_registro}.docx")
    output_pdf = os.path.join(OUTPUT_DIR, f"certificado_{numero_registro}.pdf")
    doc.save(output_word)
    
    # Convertir a PDF
    convert(output_word, output_pdf)

print("¡Certificados generados exitosamente en PDF y Word!")
